# -*- coding: utf-8 -*-
# @Time:2020/7/10 10:05
# @Pile:lesson1.py
# @QQ:12133927
# @Company:湖南省柠檬信息技术
#  Excel三大对象
# 工作簿（WorkBook)
# 表单(sheet)
# 单元格(Cell)

import requests
import openpyxl
#读取测试用例函数
def read_data(filename,sheetname):  #开始封装函数：准备定义函数，设置形参如括号里的就是形参
    wb = openpyxl.load_workbook(filename)  #定义参数
    sheet = wb[sheetname]    #定义参数
    max_row=sheet.max_row       #获取最大行数
    max_column=sheet.max_column  #获取最大列数
    case_list=[]   #创建空列表，存放测试用例
    for i in range(2,max_row+1):   #使用for循环在获取行列范围内数据信息，取头不取尾，所以+1
        dict1=dict(
        case_id=sheet.cell(row=i,column=1).value,  #获取用例编号
        url=sheet.cell(row=i,column=5).value,   #获取地址
        data=sheet.cell(row=i,column=6).value,  #获取正文数据
        expect = sheet.cell(row=i,column=7).value,  #获取预期结果)
        )
        case_list.append(dict1)  #每循环一次，就把读取到字典的数据存放进去list列表
    return case_list  #设置返回值return，返回测试用例

#执行接口函数
def api_fun(url,data):
   headers_reg={"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}#请求头，字典
   res=requests.post(url=url,json=data,headers=headers_reg)   #接收post方法的结果 自己定义res
   response=res.json()   #响应正文
   return response

#写入结果
def write_result(filename,sheetname,row,column,final_result): #把行列全部设置形参。
    wb = openpyxl.load_workbook(filename) #定义函数，定义形参。
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value=final_result  #写入结果
    wb.save(filename) #保存，要关闭文档

#执行测试用例并回写实际结果
def execute_fun(filename,sheetname):
    cases=read_data(filename,sheetname)  #调用读取测试用例，获取所有测试用例数据保存到变量,设置变量cases
    for case in cases:           #用for循环，提取数据
        case_id = case.get('case_id')  #case['case_id'] 每循环一次，都是取的每一条用例的ID
        url = case.get('url')
        data = eval(case.get('data')) #eval()运行被字符串包裹的表达式---去掉字符串引号
        expect = eval(case.get('expect') )  #获取预期结果
        expect_msg = expect.get('msg')#获取预期结果中的msg
        real_result=api_fun(url=url,data=data) #调用函数api_fun发送接口请求函数，返回结果用变量real_result接收
        real_msg = real_result.get('msg') #获取实际结果中的msg
        print('获取预期结果中的msg:{}'.format(expect_msg))
        print('获取实际结果中的msg:{}'.format(real_msg))
        if real_msg==expect_msg:
             print('第{}条测试用例通过！'.format(case_id))
             final_re='Passed'
        else:
             print('第{}条测试用例不通过！'.format(case_id))
             final_re = 'Failed'
        write_result(filename,sheetname,case_id+1,8,final_re)
        print('*'*30)
execute_fun('test_case_api.xlsx','login')









